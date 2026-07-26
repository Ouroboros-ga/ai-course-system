"""Phase B 专用 Excel 导入器

仅读取指定 Excel 文件，不遍历课件目录。
将教学问答数据集导入为 QuestionBankItem 记录，默认 status=unassigned。

Excel 结构（已核验）：
  单表、3123行、11列
  字段：规则分类、标准问题、答案、规则状态、匹配模式、6个相似问法
  无课程、知识点、课件页码或出处字段

使用方法：
  python -m app.tools.import_question_bank "D:\\path\\to\\教学问答数据集.xlsx"
"""
from __future__ import annotations

import hashlib
import os
import sys
from datetime import datetime
from typing import Any

from sqlmodel import Session, select

from app.core.time_utils import utcnow_naive

from app.models.database import engine
from app.models.question_bank_model import (
    QuestionBankItem,
    QuestionStatus,
    QuestionType,
    QuestionDifficulty,
)


# Excel 列名映射（中文 -> 内部字段）
COLUMN_MAP = {
    "规则分类": "category",
    "标准问题": "question_text",
    "答案": "answer",
    "规则状态": "rule_status",
    "匹配模式": "match_mode",
    "相似问法1": "similar_1",
    "相似问法2": "similar_2",
    "相似问法3": "similar_3",
    "相似问法4": "similar_4",
    "相似问法5": "similar_5",
    "相似问法6": "similar_6",
}

# 可能的相似问法列名变体
SIMILAR_COLUMN_VARIANTS = [
    "相似问法1", "相似问法2", "相似问法3", "相似问法4", "相似问法5", "相似问法6",
    "相似问法 1", "相似问法 2", "相似问法 3", "相似问法 4", "相似问法 5", "相似问法 6",
    "相似问题1", "相似问题2", "相似问题3", "相似问题4", "相似问题5", "相似问题6",
]
REQUIRED_COLUMNS = {"规则分类", "标准问题", "答案", "规则状态", "匹配模式"}
MAX_IMPORT_BYTES = 128 * 1024 * 1024


def _file_sha256(file_path: str) -> str:
    digest = hashlib.sha256()
    with open(file_path, "rb") as source:
        for chunk in iter(lambda: source.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def _bytes_sha256(content: bytes) -> str:
    """计算字节内容的 SHA256（用于对象存储读取后的哈希校验）"""
    return hashlib.sha256(content).hexdigest()


def _read_excel_rows(file_path: str) -> list[dict[str, Any]]:
    """读取 Excel 文件，返回字典列表。

    使用 openpyxl 读取，不依赖 pandas。
    """
    try:
        from openpyxl import load_workbook
    except ImportError:
        raise RuntimeError("需要 openpyxl: pip install openpyxl")

    wb = load_workbook(file_path, read_only=True, data_only=True)
    ws = wb.active

    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        wb.close()
        return []

    headers = [str(h).strip() if h is not None else "" for h in rows[0]]
    missing = REQUIRED_COLUMNS - set(headers)
    if missing:
        wb.close()
        raise ValueError(f"Excel 缺少必需列: {', '.join(sorted(missing))}")

    result = []
    for row in rows[1:]:
        if all(cell is None or str(cell).strip() == "" for cell in row):
            continue
        row_dict = {}
        for i, header in enumerate(headers):
            if i < len(row):
                row_dict[header] = row[i]
        result.append(row_dict)

    wb.close()
    return result


def _read_excel_rows_from_bytes(content: bytes) -> list[dict[str, Any]]:
    """从字节内容读取 Excel 行（用于对象存储读取路径）。

    与 `_read_excel_rows` 共用解析逻辑，但不依赖文件路径。
    """
    try:
        from openpyxl import load_workbook
    except ImportError:
        raise RuntimeError("需要 openpyxl: pip install openpyxl")
    import io

    wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    ws = wb.active

    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        wb.close()
        return []

    headers = [str(h).strip() if h is not None else "" for h in rows[0]]
    missing = REQUIRED_COLUMNS - set(headers)
    if missing:
        wb.close()
        raise ValueError(f"Excel 缺少必需列: {', '.join(sorted(missing))}")

    result = []
    for row in rows[1:]:
        if all(cell is None or str(cell).strip() == "" for cell in row):
            continue
        row_dict = {}
        for i, header in enumerate(headers):
            if i < len(row):
                row_dict[header] = row[i]
        result.append(row_dict)

    wb.close()
    return result


def _extract_similar_questions(row: dict[str, Any]) -> list[str]:
    """从行数据中提取所有相似问法"""
    similar = []
    for variant in SIMILAR_COLUMN_VARIANTS:
        value = row.get(variant)
        if value and str(value).strip():
            similar.append(str(value).strip())
    return similar


def _map_row_to_item(
    row: dict[str, Any],
    row_index: int,
    batch_id: str,
    *,
    course_id: int | None = None,
) -> QuestionBankItem:
    """将 Excel 行映射为 QuestionBankItem

    Args:
        row: Excel 行字典
        row_index: 行号（用于审计）
        batch_id: 导入批次ID
        course_id: 课程ID；None 表示未归属（CLI 默认），API 触发时传入具体课程
    """
    # openpyxl 对空单元格返回 None；统一转为空字符串再 strip
    def _cell_str(value: Any) -> str:
        return str(value).strip() if value is not None else ""

    question_text = _cell_str(row.get("标准问题"))
    answer = _cell_str(row.get("答案"))
    category = _cell_str(row.get("规则分类"))
    rule_status = _cell_str(row.get("规则状态"))
    match_mode = _cell_str(row.get("匹配模式"))
    similar = _extract_similar_questions(row)

    return QuestionBankItem(
        question_text=question_text,
        answer=answer,
        options={},
        similar_questions=similar,
        question_type=QuestionType.SHORT_ANSWER,
        difficulty=QuestionDifficulty.MEDIUM,
        category=category,
        match_mode=match_mode,
        rule_status=rule_status,
        course_id=course_id,  # None=未归属；API 触发时绑定到具体课程
        knowledge_node_ids=[],
        prerequisite_node_ids=[],
        status=QuestionStatus.UNASSIGNED,
        version=1,
        prev_version_id=None,
        is_latest=True,
        import_batch_id=batch_id,
        source_row_index=row_index,
        generated_by="excel_import",
        generation_metadata={},
        created_at=utcnow_naive(),
        updated_at=utcnow_naive(),
    )


def import_excel_to_question_bank(
    file_path: str,
    *,
    dry_run: bool = False,
    batch_id: str | None = None,
) -> dict[str, Any]:
    """导入 Excel 文件到题库

    Args:
        file_path: Excel 文件路径
        dry_run: 仅解析不写入数据库
        batch_id: 导入批次ID（不传则自动生成）

    Returns:
        导入统计信息
    """
    if not os.path.isfile(file_path):
        raise FileNotFoundError(f"Excel 文件不存在: {file_path}")
    if os.path.splitext(file_path)[1].lower() != ".xlsx":
        raise ValueError("仅允许导入显式指定的 .xlsx 文件")
    file_size = os.path.getsize(file_path)
    if file_size > MAX_IMPORT_BYTES:
        raise ValueError(f"Excel 文件超过 {MAX_IMPORT_BYTES // (1024 * 1024)} MiB 上限")

    source_hash = _file_sha256(file_path)

    if batch_id is None:
        # 相同原件产生相同批次身份，重复运行不会复制整批题目。
        batch_id = f"excel-sha256-{source_hash}"

    rows = _read_excel_rows(file_path)
    total_rows = len(rows)

    if dry_run:
        return {
            "batch_id": batch_id,
            "total_rows": total_rows,
            "dry_run": True,
            "imported": 0,
            "source_hash": source_hash,
        }

    session = Session(engine)
    imported = 0
    skipped = 0
    errors = []

    try:
        existing = session.exec(
            select(QuestionBankItem.id)
            .where(QuestionBankItem.import_batch_id == batch_id)
            .limit(1)
        ).first()
        if existing is not None:
            return {
                "batch_id": batch_id,
                "source_hash": source_hash,
                "total_rows": total_rows,
                "imported": 0,
                "skipped": total_rows,
                "dry_run": False,
                "already_imported": True,
                "errors": [],
            }

        for index, row in enumerate(rows, start=2):  # Excel行号从2开始(1是表头)
            question_text = str(row.get("标准问题", "")).strip()
            if not question_text:
                skipped += 1
                continue

            item = _map_row_to_item(row, index, batch_id)
            session.add(item)
            imported += 1

        session.commit()
    except Exception as e:
        session.rollback()
        errors.append(str(e))
        raise
    finally:
        session.close()

    return {
        "batch_id": batch_id,
        "source_hash": source_hash,
        "total_rows": total_rows,
        "imported": imported,
        "skipped": skipped,
        "dry_run": False,
        "already_imported": False,
        "errors": errors,
    }


def main():
    """命令行入口"""
    if len(sys.argv) < 2:
        print("用法: python -m app.tools.import_question_bank <excel路径> [--dry-run]")
        sys.exit(1)

    file_path = sys.argv[1]
    dry_run = "--dry-run" in sys.argv

    print(f"导入文件: {file_path}")
    print(f"模式: {'dry-run' if dry_run else '实际导入'}")

    result = import_excel_to_question_bank(file_path, dry_run=dry_run)
    print(f"结果: {result}")


if __name__ == "__main__":
    main()
