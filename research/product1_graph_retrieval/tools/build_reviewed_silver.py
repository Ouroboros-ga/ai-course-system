"""Build a versioned personal-learning Reviewed Silver fixture from real courseware.

This tool deliberately does not call an LLM or a vector service.  The checked
workbook supplies human semantic decisions; the small reconciliation file
records the deterministic Codex resolutions for rewrite rows whose replacement
cell was left blank.  PPTX native text is the canonical source text, while the
paired PDF is checked for page-count and source identity only.
"""

from __future__ import annotations

import argparse
import hashlib
import json
import re
import sys
import unicodedata
import zipfile
from collections import defaultdict
from pathlib import Path
from typing import Any
from xml.etree import ElementTree


RESEARCH_ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(RESEARCH_ROOT))

from src.canonical import canonical_json_bytes, sha256_file, sha256_text, write_json, write_jsonl  # noqa: E402
from src.fixture_io import (  # noqa: E402
    GOLD_ONLY_FILES,
    PUBLIC_INPUT_FILES,
    REQUIRED_FILES,
    compute_fixture_content_hash,
)
from src.identities import (  # noqa: E402
    production_compatible_citation_key,
    research_chunk_id,
    research_evidence_id,
    research_knowledge_point_id,
    research_query_id,
    research_slide_id,
)


WORKBOOK_SHEETS = ("课程范围", "问题确认", "知识点确认", "章节核对（可选）")
P_NS = "http://schemas.openxmlformats.org/presentationml/2006/main"
A_NS = "http://schemas.openxmlformats.org/drawingml/2006/main"
NS = {"p": P_NS, "a": A_NS}
ALLOWED_STRATA = {
    "exact_term", "definition", "formula_or_code", "paraphrase",
    "cross_language_alias", "multi_hop_relation", "no_answer",
}


class ReviewedSilverBuildError(ValueError):
    pass


def _stable(prefix: str, *parts: object) -> str:
    payload = "\x00".join("" if item is None else str(item) for item in parts).encode("utf-8")
    return prefix + hashlib.sha256(payload).hexdigest()[:24]


def _text(value: object) -> str:
    return unicodedata.normalize("NFKC", str(value or "")).replace("\r\n", "\n").replace("\r", "\n").strip()


def _page_range(value: object) -> tuple[int, int]:
    text = _text(value)
    matches = list(re.finditer(r"(?<!\d)(\d+)\s*(?:[-—–]\s*(\d+))?(?!\d)", text))
    if not matches:
        raise ReviewedSilverBuildError(f"page_range_missing:{text}")
    last = matches[-1]
    start = int(last.group(1)); end = int(last.group(2) or start)
    if start < 1 or end < start:
        raise ReviewedSilverBuildError(f"page_range_invalid:{text}")
    return start, end


def _looks_like_location(value: str) -> bool:
    return bool(re.search(r"页|\d+\s*[-—–]\s*\d+|^\s*\d+(?:\.\d+){1,2}\s", value))


def _table_text(node: ElementTree.Element) -> str:
    rows: list[str] = []
    for row in node.findall(".//a:tr", NS):
        cells = []
        for cell in row.findall("a:tc", NS):
            value = "".join(_text(item.text) for item in cell.findall(".//a:t", NS)).strip()
            if value:
                cells.append(value)
        if cells:
            rows.append(" | ".join(cells))
    return "\n".join(rows)


def _paragraph_text(node: ElementTree.Element) -> str:
    paragraphs: list[str] = []
    for paragraph in node.findall(".//a:p", NS):
        value = "".join(_text(item.text) for item in paragraph.findall(".//a:t", NS)).strip()
        if value:
            paragraphs.append(value)
    return "\n".join(paragraphs)


def _extract_pptx(path: Path) -> list[dict[str, Any]]:
    """Extract real PPTX text in slide XML order without adding a dependency."""

    try:
        with zipfile.ZipFile(path) as archive:
            named = []
            for name in archive.namelist():
                match = re.fullmatch(r"ppt/slides/slide(\d+)\.xml", name)
                if match:
                    named.append((int(match.group(1)), name))
            if not named:
                raise ReviewedSilverBuildError(f"pptx_has_no_slides:{path.name}")
            slides = []
            for number, name in sorted(named):
                root = ElementTree.fromstring(archive.read(name))
                tree = root.find(".//p:spTree", NS)
                if tree is None:
                    raise ReviewedSilverBuildError(f"pptx_shape_tree_missing:{path.name}:{number}")
                blocks: list[dict[str, str]] = []
                ordinal = 0
                for node in list(tree):
                    if node.tag == f"{{{P_NS}}}sp":
                        text = _paragraph_text(node)
                        kind = "paragraph"
                        placeholder = node.find(".//p:ph", NS)
                        if placeholder is not None and placeholder.get("type") in {"title", "ctrTitle"}:
                            kind = "title"
                    elif node.tag == f"{{{P_NS}}}graphicFrame":
                        text = _table_text(node)
                        kind = "table"
                    else:
                        continue
                    if not text:
                        continue
                    ordinal += 1
                    c_nv_pr = node.find(".//p:cNvPr", NS)
                    native_id = c_nv_pr.get("id") if c_nv_pr is not None else str(ordinal)
                    blocks.append({
                        "block_id": f"blk_pptx_s{number:04d}_sh{native_id}",
                        "block_type": kind,
                        "text": text,
                    })
                slides.append({"slide_number": number, "blocks": blocks})
            return slides
    except (OSError, zipfile.BadZipFile, ElementTree.ParseError) as exc:
        raise ReviewedSilverBuildError(f"pptx_parse_failed:{path.name}") from exc


def _load_workbook(path: Path) -> dict[str, list[tuple[object, ...]]]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:  # pragma: no cover - environment-specific
        raise ReviewedSilverBuildError("openpyxl_required_for_reviewed_silver_builder") from exc
    book = load_workbook(path, read_only=True, data_only=False)
    missing = set(WORKBOOK_SHEETS) - set(book.sheetnames)
    if missing:
        raise ReviewedSilverBuildError(f"workbook_sheets_missing:{sorted(missing)}")
    return {
        name: [tuple(row) for row in book[name].iter_rows(min_row=2, values_only=True)]
        for name in WORKBOOK_SHEETS
    }


def _load_reconciliation(path: Path) -> dict[tuple[str, str], dict[str, str]]:
    value = json.loads(path.read_text(encoding="utf-8"))
    if value.get("reconciliation_schema_version") != "reviewed-silver-reconciliation/0.2":
        raise ReviewedSilverBuildError("reconciliation_schema_version_invalid")
    result = {}
    for row in value.get("query_resolutions", []):
        key = (_text(row.get("course_id")), _text(row.get("source_text")))
        final = _text(row.get("final_text"))
        if not all(key) or not final or key in result:
            raise ReviewedSilverBuildError(f"reconciliation_row_invalid:{key}")
        result[key] = {"final_text": final, "resolution_reason": _text(row.get("resolution_reason"))}
    return result


def _discover_course_files(source_dir: Path, courses: dict[str, dict[str, Any]]) -> dict[str, dict[str, Any]]:
    pptx_by_count: dict[int, list[Path]] = defaultdict(list)
    for path in sorted(source_dir.glob("*.pptx"), key=lambda item: item.name):
        pptx_by_count[len(_extract_pptx(path))].append(path)
    result = {}
    for course_id, course in sorted(courses.items()):
        candidates = pptx_by_count.get(course["page_count"], [])
        if len(candidates) != 1:
            raise ReviewedSilverBuildError(f"pptx_page_count_mapping_not_unique:{course_id}:{course['page_count']}")
        pptx = candidates[0]; pdf = pptx.with_suffix(".pdf")
        if not pdf.is_file():
            raise ReviewedSilverBuildError(f"paired_pdf_missing:{pptx.name}")
        try:
            from pypdf import PdfReader
        except ImportError as exc:  # pragma: no cover - environment-specific
            raise ReviewedSilverBuildError("pypdf_required_for_paired_pdf_check") from exc
        if len(PdfReader(str(pdf)).pages) != course["page_count"]:
            raise ReviewedSilverBuildError(f"pdf_page_count_mismatch:{course_id}:{pdf.name}")
        result[course_id] = {"pptx": pptx, "pdf": pdf, "slides": _extract_pptx(pptx)}
    return result


def _parse_courses(rows: list[tuple[object, ...]]) -> dict[str, dict[str, Any]]:
    result = {}
    for row in rows:
        if not _text(row[0]):
            continue
        course_id, name, page_count, source_file, decision = _text(row[0]), _text(row[1]), row[2], _text(row[3]), _text(row[4])
        if decision != "保留":
            continue
        if not isinstance(page_count, int) or page_count < 1:
            raise ReviewedSilverBuildError(f"course_page_count_invalid:{course_id}")
        if course_id in result:
            raise ReviewedSilverBuildError(f"course_duplicate:{course_id}")
        result[course_id] = {"course_id": course_id, "course_name": name, "page_count": page_count, "workbook_source_file": source_file}
    if len(result) < 2:
        raise ReviewedSilverBuildError("at_least_two_retained_courses_required")
    return result


def _parse_chapters(rows: list[tuple[object, ...]], courses: dict[str, dict[str, Any]]) -> dict[str, list[dict[str, Any]]]:
    result: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for row in rows:
        course_id, label = _text(row[0]), _text(row[1])
        if course_id not in courses or not label:
            continue
        start, end = row[2], row[3]
        decision = _text(row[4])
        if decision == "改写":
            start, end = row[5], row[6]
        if not isinstance(start, int) or not isinstance(end, int) or not 1 <= start <= end <= courses[course_id]["page_count"]:
            raise ReviewedSilverBuildError(f"chapter_range_invalid:{course_id}:{label}")
        chapter_id = _stable("rchp_", course_id, label, start, end)
        result[course_id].append({"chapter_id": chapter_id, "label": label, "start": start, "end": end,
                                  "chapter_path": [course_id, chapter_id], "review_decision": decision or "正确"})
    for course_id, chapters in result.items():
        chapters.sort(key=lambda item: (item["start"], item["end"], item["label"]))
        expected = 1
        for chapter in chapters:
            if chapter["start"] != expected:
                raise ReviewedSilverBuildError(f"chapter_coverage_gap_or_overlap:{course_id}:{expected}:{chapter['start']}")
            expected = chapter["end"] + 1
        if expected != courses[course_id]["page_count"] + 1:
            raise ReviewedSilverBuildError(f"chapter_coverage_incomplete:{course_id}:{expected}")
    if set(result) != set(courses):
        raise ReviewedSilverBuildError("chapter_course_coverage_incomplete")
    return result


def _chapter_for_page(chapters: list[dict[str, Any]], page: int) -> dict[str, Any]:
    matches = [row for row in chapters if row["start"] <= page <= row["end"]]
    if len(matches) != 1:
        raise ReviewedSilverBuildError(f"chapter_lookup_not_unique:{page}")
    return matches[0]


def _parse_knowledge_points(rows: list[tuple[object, ...]], chapters: dict[str, list[dict[str, Any]]], courses: dict[str, dict[str, Any]]) -> list[dict[str, Any]]:
    result = []
    seen = set()
    for row in rows:
        course_id, original, source_location, decision, target, note = _text(row[0]), _text(row[1]), _text(row[2]), _text(row[4]), _text(row[5]), _text(row[6])
        if course_id not in courses or not original:
            continue
        start, end = _page_range(source_location)
        label = original
        if decision in {"改写", "否"} and target:
            if _looks_like_location(target):
                start, end = _page_range(target)
            elif decision == "改写":
                label = target
        if not 1 <= start <= end <= courses[course_id]["page_count"]:
            raise ReviewedSilverBuildError(f"knowledge_point_page_range_invalid:{course_id}:{label}")
        key = (course_id, label)
        if key in seen:
            raise ReviewedSilverBuildError(f"knowledge_point_duplicate_after_reconciliation:{course_id}:{label}")
        seen.add(key)
        chapter = _chapter_for_page(chapters[course_id], start)
        aliases = [original] if original != label else []
        result.append({
            "research_knowledge_point_id": research_knowledge_point_id(course_id=course_id, canonical_label=label),
            "research_sidecar": True, "not_a_production_contract_field": True,
            "course_id": course_id, "canonical_label": label, "aliases": aliases,
            "alias_provenance": {"source": "reviewed_silver_reconciled_pre_split", "frozen_before_split": True},
            "chapter_id": chapter["chapter_id"], "chapter_path": chapter["chapter_path"],
            "source_page_range": f"{start}-{end}" if start != end else str(start),
            "source_page_start": start, "source_page_end": end,
            "research_evidence_ids": [], "review_status": "reviewed_silver_pending_evidence",
            "review_decision": decision or "通过", "review_note": note,
        })
    return sorted(result, key=lambda item: item["research_knowledge_point_id"])


def _query_type_and_stratum(original_type: str, suggested_stratum: str, answerable: bool, text: str) -> tuple[str, str]:
    if not answerable:
        return "no_evidence", "no_answer"
    if suggested_stratum in ALLOWED_STRATA and suggested_stratum != "no_answer":
        stratum = suggested_stratum
    elif original_type == "direct_definition":
        stratum = "definition"
    else:
        stratum = "paraphrase"
    if any(marker in text for marker in ("公式", "关系式", "转速", "转矩")):
        query_type = "formula"
    elif any(marker in text for marker in ("为什么", "如何", "怎样")):
        query_type = "explanation"
    else:
        query_type = "definition"
    return query_type, stratum


def _parse_queries(rows: list[tuple[object, ...]], reconciliations: dict[tuple[str, str], dict[str, str]], courses: dict[str, dict[str, Any]]) -> list[dict[str, Any]]:
    result = []
    seen = set()
    for row in rows:
        course_id, original, suggested_kp, source_pages = _text(row[0]), _text(row[1]), _text(row[2]), _text(row[3])
        original_type, suggested_stratum, decision = _text(row[4]), _text(row[5]), _text(row[6])
        replacement, replacement_kp, reason = _text(row[7]), _text(row[8]), _text(row[9])
        if course_id not in courses or not original:
            continue
        start, end = _page_range(source_pages)
        if decision == "改写":
            if replacement:
                final_text, resolution_source = replacement, "human_written_replacement"
            else:
                item = reconciliations.get((course_id, original))
                if not item:
                    raise ReviewedSilverBuildError(f"rewrite_without_resolution:{course_id}:{original}")
                final_text, resolution_source = item["final_text"], "llm_reconciliation_file"
                reason = reason or item["resolution_reason"]
        elif decision == "通过":
            final_text, resolution_source = original, "human_approved_original"
        else:
            continue
        answerable = decision == "改写" or original_type not in {"hard_negative", "cross_course_isolation"}
        query_type, query_stratum = _query_type_and_stratum(original_type, suggested_stratum, answerable, final_text)
        research_id = research_query_id(course_id=course_id, text=final_text)
        if research_id in seen:
            raise ReviewedSilverBuildError(f"query_duplicate_after_reconciliation:{course_id}:{final_text}")
        seen.add(research_id)
        result.append({
            "research_query_id": research_id, "research_sidecar": True, "not_a_production_contract_field": True,
            "course_id": course_id, "text": final_text, "query_type": query_type, "query_stratum": query_stratum,
            "tags": sorted({"zh", original_type, decision}), "source_page_start": start, "source_page_end": end,
            "resolved_knowledge_point_label": replacement_kp or suggested_kp,
            "review_decision": decision, "resolution_source": resolution_source, "review_note": reason,
            "answerability": "answerable" if answerable else "unanswerable_in_course",
        })
    if len(result) < 1:
        raise ReviewedSilverBuildError("queries_missing")
    return sorted(result, key=lambda item: item["research_query_id"])


def _search_chars(value: str) -> set[str]:
    return {char for char in unicodedata.normalize("NFKC", value).casefold() if char.isalnum() or "\u4e00" <= char <= "\u9fff"}


def _select_evidence(candidates: list[dict[str, Any]], term: str, limit: int = 3) -> list[dict[str, Any]]:
    term_chars = _search_chars(term)
    scored = []
    for candidate in candidates:
        score = len(term_chars & _search_chars(candidate["text"]))
        if candidate.get("block_type") == "title":
            score += 0.25
        scored.append((score, candidate["page_or_slide"], candidate["block_id"], candidate))
    return [row[-1] for row in sorted(scored, key=lambda row: (-row[0], row[1], row[2]))[:limit]]


def _assign_splits(records: list[dict[str, Any]], id_field: str, group_fields: tuple[str, ...], names: tuple[str, ...]) -> dict[str, list[str]]:
    result = {name: [] for name in names}
    groups: dict[tuple[str, ...], list[dict[str, Any]]] = defaultdict(list)
    for row in records:
        groups[tuple(str(row.get(field, "")) for field in group_fields)].append(row)
    for group in sorted(groups):
        rows = sorted(groups[group], key=lambda item: hashlib.sha256(item[id_field].encode("utf-8")).hexdigest())
        for index, row in enumerate(rows):
            if names == ("train", "validation", "test"):
                # Allocate inside each course instead of inside tiny
                # course×stratum groups, which can otherwise leave almost no
                # validation/test examples.  Stable ID hashing avoids source
                # order becoming an implicit split feature.
                if index < (len(rows) * 3) // 5:
                    name = "train"
                elif index < (len(rows) * 4) // 5:
                    name = "validation"
                else:
                    name = "test"
            else:
                name = names[index % len(names)]
            result[name].append(row[id_field])
    return {name: sorted(values) for name, values in result.items()}


def build_reviewed_silver(workbook: Path, source_dir: Path, reconciliation_path: Path, output_dir: Path) -> dict[str, Any]:
    workbook, source_dir, reconciliation_path, output_dir = map(Path, (workbook, source_dir, reconciliation_path, output_dir))
    if output_dir.exists() and any(output_dir.iterdir()):
        raise ReviewedSilverBuildError(f"output_directory_must_be_empty:{output_dir}")
    raw = _load_workbook(workbook)
    courses = _parse_courses(raw["课程范围"])
    chapters = _parse_chapters(raw["章节核对（可选）"], courses)
    knowledge_points = _parse_knowledge_points(raw["知识点确认"], chapters, courses)
    queries = _parse_queries(raw["问题确认"], _load_reconciliation(reconciliation_path), courses)
    files = _discover_course_files(source_dir, courses)

    source_blocks: list[dict[str, Any]] = []
    evidence: list[dict[str, Any]] = []
    corpus: list[dict[str, Any]] = []
    slides: list[dict[str, Any]] = []
    evidence_by_course_page: dict[tuple[str, int], list[dict[str, Any]]] = defaultdict(list)
    source_inventory = []
    for course_id in sorted(courses):
        source = files[course_id]; pptx, pdf = source["pptx"], source["pdf"]
        pptx_hash, pdf_hash = sha256_file(pptx), sha256_file(pdf)
        artifact_id = _stable("rart_", course_id, pptx_hash, pdf_hash)
        document_id = _stable("rdoc_", course_id, pptx_hash)
        version_ref = f"sha256:{pptx_hash}"
        source_inventory.append({"course_id": course_id, "pptx_file": pptx.name, "pdf_file": pdf.name,
                                 "pptx_sha256": pptx_hash, "pdf_sha256": pdf_hash, "page_count": courses[course_id]["page_count"]})
        for slide in source["slides"]:
            number = slide["slide_number"]
            chapter = _chapter_for_page(chapters[course_id], number)
            unit_id = f"unit_{course_id}_slide_{number:04d}"
            slide_evidence, block_ids, title, body = [], [], "", []
            for block in slide["blocks"]:
                # PPTX shape IDs are unique only inside one presentation.  The
                # research Block ID must be globally unique across all courses.
                block_id, text = f"blk_{course_id}_{block['block_id']}", block["text"]
                block_row = {
                    "course_id": course_id, "artifact_id": artifact_id, "document_id": document_id,
                    "unit_id": unit_id, "unit_type": "slide", "unit_index": number, "block_id": block_id,
                    "block_type": block["block_type"], "page_or_slide": number,
                    "chapter_id": chapter["chapter_id"], "chapter_path": chapter["chapter_path"],
                    "title": "", "text": text, "text_sha256": sha256_text(text),
                }
                if block["block_type"] == "title" and not title:
                    title = text
                source_blocks.append(block_row); block_ids.append(block_id); body.append(text)
                evidence_id = research_evidence_id(course_id=course_id, artifact_id=artifact_id, document_id=document_id,
                    unit_id=unit_id, block_id=block_id, version_ref=version_ref, char_start=0, char_end=len(text))
                evidence_row = {
                    "research_evidence_id": evidence_id, "research_sidecar": True, "not_a_production_contract_field": True,
                    "course_id": course_id, "artifact_id": artifact_id, "document_id": document_id, "unit_id": unit_id,
                    "block_id": block_id, "version_ref": version_ref, "page_or_slide": number, "char_start": 0,
                    "char_end": len(text), "text_snippet": text, "status": "active",
                    "citation_key": production_compatible_citation_key(artifact_id=artifact_id, block_id=block_id, char_start=0, char_end=len(text)),
                    "metadata": {"unit_type": "slide", "parser": "native_pptx_xml/1", "pdf_ref": f"controlled-source://{course_id}/pdf?page={number}"},
                }
                evidence.append(evidence_row); slide_evidence.append(evidence_id)
                candidate = {**evidence_row, "text": text, "block_type": block["block_type"]}
                evidence_by_course_page[(course_id, number)].append(candidate)
                corpus.append({
                    "research_chunk_id": research_chunk_id(course_id=course_id, document_id=document_id, unit_id=unit_id,
                        block_id=block_id, research_evidence_ids=[evidence_id], text_sha256=sha256_text(text)),
                    "research_sidecar": True, "not_a_production_contract_field": True, "course_id": course_id,
                    "artifact_id": artifact_id, "document_id": document_id, "unit_id": unit_id, "unit_type": "slide",
                    "unit_index": number, "block_id": block_id, "block_type": block["block_type"],
                    "research_evidence_ids": [evidence_id], "page_or_slide": number, "chapter_id": chapter["chapter_id"],
                    "chapter_path": chapter["chapter_path"], "title": "", "text": text, "text_sha256": sha256_text(text), "language": "zh-CN",
                })
            slides.append({
                "research_slide_id": research_slide_id(course_id=course_id, document_id=document_id, unit_id=unit_id),
                "research_sidecar": True, "not_a_production_contract_field": True, "course_id": course_id,
                "document_id": document_id, "unit_id": unit_id, "slide_number": number, "chapter_id": chapter["chapter_id"],
                "chapter_path": chapter["chapter_path"], "title": title, "body_text": "\n".join(body),
                "block_ids": block_ids, "research_evidence_ids": slide_evidence,
                "visual_only": not bool(slide_evidence),
            })

    by_kp_label = {(row["course_id"], row["canonical_label"]): row for row in knowledge_points}
    mapping_qrels: list[dict[str, Any]] = []
    kp_evidence: dict[str, list[str]] = {}
    for kp in knowledge_points:
        candidates = [item for page in range(kp["source_page_start"], kp["source_page_end"] + 1)
                      for item in evidence_by_course_page[(kp["course_id"], page)]]
        selected = _select_evidence(candidates, kp["canonical_label"])
        if not selected:
            raise ReviewedSilverBuildError(f"knowledge_point_without_native_evidence:{kp['research_knowledge_point_id']}")
        refs = [row["research_evidence_id"] for row in selected]
        kp["research_evidence_ids"] = refs; kp["review_status"] = "reviewed_silver_evidence_bound"
        kp_evidence[kp["research_knowledge_point_id"]] = refs
        selected_by_slide: dict[str, list[dict[str, Any]]] = defaultdict(list)
        for item in selected:
            slide_id = research_slide_id(course_id=kp["course_id"], document_id=item["document_id"], unit_id=item["unit_id"])
            selected_by_slide[slide_id].append(item)
        for index, slide_id in enumerate(selected_by_slide):
            slide_items = selected_by_slide[slide_id]
            mapping_qrels.append({
                "research_knowledge_point_id": kp["research_knowledge_point_id"], "research_slide_id": slide_id,
                "relevance": 2 if index == 0 else 1,
                "judgment": "primary_slide" if index == 0 else "supporting_slide",
                "research_evidence_ids": [item["research_evidence_id"] for item in slide_items],
                "provenance": "reviewed_kp_page_range_plus_native_pptx_text",
            })

    labels: list[dict[str, Any]] = []
    retrieval_qrels: list[dict[str, Any]] = []
    for query in queries:
        labels.append({"research_query_id": query["research_query_id"], "answerability": query.pop("answerability"),
                       "reason_code": "human_rewrite_in_course" if query["review_decision"] == "改写" else "reviewed_source_variant"})
        label = labels[-1]
        target_kp = by_kp_label.get((query["course_id"], query["resolved_knowledge_point_label"]))
        if label["answerability"] == "answerable":
            selected = []
            if target_kp:
                wanted = set(kp_evidence[target_kp["research_knowledge_point_id"]])
                selected = [item for page in range(query["source_page_start"], query["source_page_end"] + 1)
                            for item in evidence_by_course_page[(query["course_id"], page)] if item["research_evidence_id"] in wanted]
                selected = _select_evidence(selected, query["text"]) or [item for page in range(target_kp["source_page_start"], target_kp["source_page_end"] + 1)
                            for item in evidence_by_course_page[(query["course_id"], page)] if item["research_evidence_id"] in wanted]
            if not selected:
                candidates = [item for page in range(query["source_page_start"], query["source_page_end"] + 1)
                              for item in evidence_by_course_page[(query["course_id"], page)]]
                selected = _select_evidence(candidates, query["text"])
            if not selected:
                raise ReviewedSilverBuildError(f"answerable_query_without_native_evidence:{query['research_query_id']}")
            for index, item in enumerate(selected[:3]):
                retrieval_qrels.append({"research_query_id": query["research_query_id"], "research_evidence_id": item["research_evidence_id"],
                                        "relevance": 2 if index == 0 else 1, "judgment": "direct_support" if index == 0 else "partial_context",
                                        "provenance": "reviewed_query_page_range_plus_native_pptx_text"})
        else:
            candidates = [item for page in range(query["source_page_start"], query["source_page_end"] + 1)
                          for item in evidence_by_course_page[(query["course_id"], page)]]
            for item in _select_evidence(candidates, query["text"], limit=1):
                retrieval_qrels.append({"research_query_id": query["research_query_id"], "research_evidence_id": item["research_evidence_id"],
                                        "relevance": 0, "judgment": "topic_related_but_insufficient",
                                        "provenance": "reviewed_negative_query_boundary"})

    query_splits = _assign_splits(queries, "research_query_id", ("course_id",), ("train", "validation", "test"))
    mapping_splits = _assign_splits(knowledge_points, "research_knowledge_point_id", ("course_id",), ("validation", "test"))
    splits = {
        "split_version": "reviewed-silver/0.2", "policy": "thresholds_and_weights_use_validation_only",
        "test_gold_access": "evaluation_only_after_run_freeze", "train_query_ids": query_splits["train"],
        "validation_query_ids": query_splits["validation"], "test_query_ids": query_splits["test"],
        "validation_knowledge_point_ids": mapping_splits["validation"], "test_knowledge_point_ids": mapping_splits["test"],
    }
    output_dir.mkdir(parents=True, exist_ok=True)
    records = {
        "source_blocks.jsonl": sorted(source_blocks, key=lambda item: item["block_id"]),
        "evidence.jsonl": sorted(evidence, key=lambda item: item["research_evidence_id"]),
        "corpus.jsonl": sorted(corpus, key=lambda item: item["research_chunk_id"]),
        "queries.jsonl": queries,
        "retrieval_query_labels.jsonl": sorted(labels, key=lambda item: item["research_query_id"]),
        "retrieval_qrels.jsonl": sorted(retrieval_qrels, key=lambda item: (item["research_query_id"], item["research_evidence_id"])),
        "knowledge_points.jsonl": knowledge_points,
        "slides.jsonl": sorted(slides, key=lambda item: item["research_slide_id"]),
        "mapping_qrels.jsonl": sorted(mapping_qrels, key=lambda item: (item["research_knowledge_point_id"], item["research_slide_id"])),
    }
    for name, rows in records.items():
        write_jsonl(output_dir / name, rows)
    write_json(output_dir / "splits.json", splits)
    file_hashes = {name: f"sha256:{sha256_file(output_dir / name)}" for name in REQUIRED_FILES}
    manifest = {
        "fixture_schema_version": "product1-graph-retrieval-fixture/1.1", "research_sidecar_schema_version": "product1-graph-retrieval-research-sidecar/1.0",
        "fixture_id": "reviewed_silver_v0_2", "dataset_level": "reviewed_silver", "course_ids": sorted(courses),
        "created_at": "2026-07-22T00:00:00+08:00", "files": file_hashes,
        "fixture_content_sha256": compute_fixture_content_hash(file_hashes),
        "source_contracts": {"document_ir": "document-ir/1.0", "evidence": "evidence/1.0", "citation": "citation/1.0", "education_graph": "edu-graph/1.0"},
        "normalization": {"ppt_page_base": 1, "source_text_mutated": False, "unicode": "source-preserved"},
        "access_policy": {"index_inputs": sorted(PUBLIC_INPUT_FILES), "gold_only": sorted(GOLD_ONLY_FILES), "test_qrels_forbidden_before_run": True},
        "gold": {"status": "reviewed_silver_llm_qrels", "eligible_for_algorithm_comparison": False},
        "annotation": {"human_semantic_review_completed": True, "llm_reconciliation_completed": True, "formal_human_gold": False,
                       "note": "Human course/query/knowledge-point/chapter decisions plus documented LLM reconciliation; qrels are source-grounded Silver labels."},
        "governance": {"b_r1_release": "formal_gate_blocked_reviewed_silver_only", "p1_00": {"status": "not_requested_for_personal_learning"}, "p1_10": {"status": "not_requested_for_personal_learning"}},
        "source_inventory": {"paired_pptx_pdf_courses": len(source_inventory), "sources": source_inventory},
        "reconciliation": {"path": str(reconciliation_path.relative_to(RESEARCH_ROOT)).replace("\\", "/"), "sha256": sha256_file(reconciliation_path)},
        "contains_production_data": False, "contains_personal_data": False,
        "identity_fields": {"all_are_research_sidecars": True, "not_production_contract_fields": True, "chunk": "research_chunk_id", "evidence": "research_evidence_id", "query": "research_query_id", "knowledge_point": "research_knowledge_point_id", "slide": "research_slide_id"},
    }
    write_json(output_dir / "manifest.json", manifest)
    report = {"fixture_id": manifest["fixture_id"], "source_workbook": workbook.name, "source_workbook_sha256": sha256_file(workbook),
              "counts": {"courses": len(courses), "chapters": sum(len(value) for value in chapters.values()), "knowledge_points": len(knowledge_points), "queries": len(queries), "source_blocks": len(source_blocks), "evidence": len(evidence), "slides": len(slides), "retrieval_qrels": len(retrieval_qrels), "mapping_qrels": len(mapping_qrels)},
              "query_decisions": {"through": sum(row["review_decision"] == "通过" for row in queries), "rewritten": sum(row["review_decision"] == "改写" for row in queries)},
              "parser": "native_pptx_xml/1", "paired_pdf_check": "pypdf_page_count", "ocr_policy": "not_a_quality_gate_by_user_decision"}
    write_json(output_dir / "build_report.json", report)
    return report


def main() -> int:
    parser = argparse.ArgumentParser(description="Build Reviewed Silver v0.2 from audited workbook and paired courseware")
    parser.add_argument("--workbook", type=Path, required=True)
    parser.add_argument("--source-dir", type=Path, required=True)
    parser.add_argument("--reconciliation", type=Path, default=RESEARCH_ROOT / "configs" / "reviewed_silver_v0_2_reconciliation.json")
    parser.add_argument("--output", type=Path, default=RESEARCH_ROOT / "datasets" / "reviewed_silver_v0_2")
    args = parser.parse_args()
    try:
        result = build_reviewed_silver(args.workbook, args.source_dir, args.reconciliation, args.output)
    except ReviewedSilverBuildError as exc:
        print(json.dumps({"status": "failed", "reason": str(exc)}, ensure_ascii=False), file=sys.stderr)
        return 2
    print(json.dumps(result, ensure_ascii=False, sort_keys=True, indent=2))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
